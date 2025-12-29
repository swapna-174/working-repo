import pandas as pd
import urllib
import sys
import time
from sqlalchemy import create_engine
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment , Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

try:
# Define database connection (replace with your connection details)
    stop_process =  0
    stop_reason = ''
##    Excel_File_Path = sys.argv[1]
##    print(Excel_File_Path)
##    time.sleep(10)
##    Excel_File_Path = r"\\Rgcedepictpi01v\d$\Census_Discharge_Reports\POC\Outbound\fax={8335960339}RecipientCompany={Aetna}_CensusReport{Cohen$#$Children$!$s$#$Medical$#$Center}_20250910.xlsx"
##    Excel_File_Path = Excel_File_Path.replace("$#$"," ")
##    Excel_File_Path = Excel_File_Path.replace("$!$","'")
  
    if len(sys.argv) > 1:
        Excel_File_Path = sys.argv[1]
        Excel_File_Path = Excel_File_Path.replace("$#$"," ")
        Excel_File_Path = Excel_File_Path.replace("$!$","'")
        ##Excel_File_Name = sys.argv[2]
        
    else:
        stop_process =  1
        stop_reason = 'No input variable provided'
        raise Exception(stop_reason)

    print(Excel_File_Path)
    if stop_process == 0:
        driver_name = 'ODBC Driver 17 for SQL Server'
        server_name = 'nonprd-cdw.et1361.epichosted.com'
        database_name = 'POC_CDW'
        
        Excel_Sheet_Name = 'Data'
        Excel_Full_File_path = Excel_File_Path 
        conn_str = (
                    f"DRIVER={{{driver_name}}};"
                    f"SERVER={server_name};"
                    f"DATABASE={database_name};"
                    f"Trusted_Connection=yes;"
                )
        conn_url = f"mssql+pyodbc:///?odbc_connect={urllib.parse.quote_plus(conn_str)}"
        engine = create_engine(conn_url, connect_args={"connect_timeout": 30})        

        # Query to fetch only the LocationName
        location_query = "SELECT DISTINCT LocationName FROM [NHX].[CensusRevLocationByPayorData]"
        location_df = pd.read_sql(location_query, engine)

        # Retrieve single LocationName if available
        if not location_df.empty:
            single_location_name = location_df['LocationName'].iloc[0]        
        else:
            stop_process = 1
            
        # Query to fetch data
        query = f"""SELECT	CONCAT('''',HospitalAccountEpicId)	AS HspAcct#
                        ,PatientName	AS PtName
                        ,DOB	
                        ,PayorName	 
                        ,SubscriberNumber	AS Subscriber#
                        ,AdmissionDate	AS AdmitDt
                        ,DXCode	AS PrimaryDX
                        ,AdmittingProvider	AS AdmitProvider
                        ,PatientClass	AS PtClass
                        ,Unit			
                        ,RoomName
        FROM [NHX].[CensusRevLocationByPayorData]"""  # Modify as needed

        # Read data into DataFrame
        df = pd.read_sql(query, engine)        

        # Create a new Excel Workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = Excel_Sheet_Name

        # Set the page layout to Landscape
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        # Write LocationName to the first row, spanning all columns
        total_columns = len(df.columns)
        location_cell = ws.cell(row=1, column=1, value=single_location_name)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)  # Merge cells
        location_cell.alignment = Alignment(horizontal='center', vertical='center')  # Center alignment
        location_cell.font = Font(name='Aptos Narrow', size=12, bold=True)

         # Write headers to the second row (excluding LocationName)
        for c_idx, col in enumerate(df.columns, 1):
            header_cell = ws.cell(row=2, column=c_idx, value=col)
            header_cell.font = Font(name='Aptos Narrow', size=9, bold=True)
            header_cell.alignment = Alignment(horizontal='left', vertical='bottom', wrap_text=False)
        
         # Define a list of columns that should not have wrapping
        no_wrap_columns = ['HspAcct#', 'DOB', 'Subscriber#', 'AdmitDt', 'PrimaryDX', 'PtClass', 'Unit']

        # Write data to the third row onwards
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 3):  # Start from row 3
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.font = Font(name='Aptos Narrow', size=9)

                # Check if the column is in the no-wrap list
                if df.columns[c_idx - 1] in no_wrap_columns:
                    cell.alignment = Alignment(horizontal='left', vertical='bottom', wrap_text=False)
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='bottom', wrap_text=True)

        # Adjust column widths
        for c_idx in range(1, len(df.columns) + 1):
            column_letter = get_column_letter(c_idx)

            # Set specific width multipliers for no-wrap columns
            if df.columns[c_idx - 1] in ['DOB']:
                header_cell = ws.cell(row=2, column=c_idx)  # Reference to header cell
                column_length = len(header_cell.value) * 3 if header_cell.value is not None else 0
            elif df.columns[c_idx - 1] in no_wrap_columns:
                # For other no-wrap columns, use 1.5x the header length
                header_cell = ws.cell(row=2, column=c_idx)  # Reference to header cell
                column_length = len(header_cell.value) * 1.5 if header_cell.value is not None else 0
            else:
                # For wrapped columns, continue using the header length
                header_cell = ws.cell(row=2, column=c_idx)  # Reference to header cell
                column_length = len(header_cell.value) if header_cell.value is not None else 0

            # Set the column width
            ws.column_dimensions[column_letter].width = column_length + 1  # Add some padding
             

        # Set print options to fit all columns on one page
        ##ws.page_setup.fitToWidth = 1  # Fit to one page width
        ##ws.page_setup.fitToHeight = 0  # Do not fit to height; can adjust if needed

        # Save the workbook to an Excel file
        wb.save(Excel_Full_File_path)
    else:
        stop_process =  1
except Exception as ex:
    print(f"Python script failed with error message:- {str(ex)}")    
